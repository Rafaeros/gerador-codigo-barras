"""
Gerador de PDF com códigos de barras dos endereços do almox em páginas separadas
"""

import os
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.graphics.barcode import code128
import openpyxl
import win32print
import win32api
from rich import print as pt
from rich.table import Table


def print_pdf(file_path: str = "enderecos.pdf"):
    printer_name = win32print.GetDefaultPrinter()
    printer = win32print.OpenPrinter(printer_name)

    abs_path = os.path.abspath(file_path)
    try:
        win32print.StartDocPrinter(printer, 1, abs_path)
        win32print.StartPagePrinter(printer)
        win32print.WritePrinter(printer, open(abs_path, "rb").read())
        win32print.EndPagePrinter(printer)
        win32print.EndDocPrinter(printer)
    finally:
        win32print.ClosePrinter(printer)


def draw_desc_barcode(c, barcode_value, x=0, y=15 * mm, max_width=95 * mm):
    """
    Desenha o código de barras em um PDF, ajustando a escala para caber no espaço disponível.
    """
    # Criação do código de barras
    barcode = code128.Code128(
        barcode_value,
        barWidth=0.5 * mm,
        barHeight=15 * mm,
        humanReadable=True,
    )

    # Largura original do código de barras
    barcode_width = barcode.width  # Largura do código de barras original

    # Calcular a escala para caber na largura máxima
    scale = min(max_width / barcode_width, 1)

    # Desenha o código de barras com escala ajustada
    c.saveState()
    c.translate(x, y)  # Posicionar no local correto
    c.scale(scale, 1)  # Ajustar apenas a largura
    barcode.drawOn(c, 0, 0)  # Desenhar no canvas
    c.restoreState()


def gerar_pdf_multiplas_paginas(enderecos: list[str]):
    """
    Gera um PDF com códigos de barras em páginas separadas.
    """
    # Dimensões do PDF em mm
    largura_pdf = 100 * mm
    altura_pdf = 45 * mm

    # Criar o PDF
    pdf = canvas.Canvas("enderecos.pdf", pagesize=(largura_pdf, altura_pdf))
    for endereco in enderecos:
        draw_desc_barcode(pdf, endereco, x=5 * mm, y=15 * mm)  # Margem ajustada
        pdf.showPage()

    # Salvar o PDF
    pdf.save()
    print("PDF gerado: enderecos.pdf")


# Exemplo de uso
def get_dados(option: int) -> list[str]:
    """
    Retorna uma lista com os dados dos códigos de barras
    """
    wb = openpyxl.load_workbook("MAPA ALMOX 2024.xlsm")
    ws = wb["MAPA ALMOX"]
    barcode_data: list[str] = []
    value: str = ""
    match option:
        # Por endereco
        case 1:
            letter: str = input(str("Digite a letra do endereco: "))
            for i, cell in enumerate(ws["F"]):
                if i == 0:
                    continue
                value = cell.value
                if value is not None:
                    if value.split("/")[1] == letter:
                        barcode_data.append(str(value))

            return barcode_data
        # Por Codigo
        case 2:
            code: str = input(str("Digite o codigo do material: "))
            for cell in ws["F"]:
                value = cell.value
                if value is not None:
                    if value.split("/")[0] == code:
                        barcode_data.append(str(value))
            return barcode_data

        # Toda a lista
        case 3:
            for i, cell in enumerate(ws["F"]):
                if i == 0:
                    continue
                value = cell.value
                if value is not None:
                    barcode_data.append(str(value))

            return barcode_data

def main():
    table = Table(title="Opcoes de Impressão")
    table.add_column("[bold underline cyan]Opção", width=5, justify="center")
    table.add_column("[bold underline cyan]Descrição")
    table.add_row("[blue]1", "Por Endereco")
    table.add_row("[blue]2", "Por Codigo do Material")
    table.add_row("[blue]3", "Planilha inteira")
    pt(table)

    option: int = int(input(str("Escolha uma opção: ")))
    barcode_data: list[str] = get_dados(option)
    print(barcode_data)
    gerar_pdf_multiplas_paginas(barcode_data)


if __name__ == "__main__":
    main()
