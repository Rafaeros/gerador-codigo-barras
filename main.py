"""
Gerador de PDF com códigos de barras dos endereços do almox em páginas separadas
"""

import os
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.graphics.barcode import code128
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
from rich import print as pt
from rich.table import Table
import win32print
#import win32api


WIDTH = 100 * mm
HEIGHT = 45 * mm

def print_pdf(file_path: str = "enderecos.pdf"):
    """
    Print the PDF file
    """
    printer_name = win32print.GetDefaultPrinter()
    printer = win32print.OpenPrinter(printer_name)

    abs_path = os.path.abspath(file_path)
    try:
        win32print.StartDocPrinter(printer, 1, abs_path)
        win32print.StartPagePrinter(printer)
        win32print.WritePrinter(printer, open(abs_path, "rb").read())
        win32print.EndPagePrinter(printer)
        win32print.EndDocPrinter(printer)
    finally:
        win32print.ClosePrinter(printer)

    #win32api.ShellExecute(0, "print", file_path, None, ".", 0)


def draw_desc_barcode(c, barcode_value, x=-10 * mm, y=15 * mm, max_width=100 * mm):
    """
    Draws the barcode on a PDF, adjusting the scale to fit the available space.
    """
    c.setFont("Arial", 13)
    barcode = code128.Code128(
        barcode_value,
        barWidth=0.30 * mm,
        barHeight=25 * mm,
        humanReadable=True,
    )
    barcode.fontName = "Arial"
    barcode.fontSize = 13
    barcode_width = barcode.width
    scale = min(max_width / barcode_width, 1)
    c.saveState()
    c.translate(x, y)
    c.scale(scale, 1)
    barcode.drawOn(c, 0, 0)
    c.restoreState()


def generate_pdf(barcode_data: list[str]):
    """
    Generates a PDF with barcodes on separate pages.
    """
    pdf = canvas.Canvas("enderecos.pdf", pagesize=(WIDTH, HEIGHT))
    pdfmetrics.registerFont(TTFont('Arial', './assets/fonts/Arial.ttf'))
    pdf.setFont("Arial", 13)
    for barcode in barcode_data:
        draw_desc_barcode(pdf, barcode, x=5 * mm, y=15 * mm)
        pdf.setFont("Arial", 16)
        pdf.drawString(85*mm, 5*mm, "PÇS")
        pdf.showPage()

    pdf.save()
    pt("[green]PDF gerado: enderecos.pdf")

# Exemplo de uso
def get_barcode_data(option: int) -> list[str]:
    """
    Returns a list with barcode data
    :return list[str]
    """
    wb = openpyxl.load_workbook("MAPA ALMOX 2024.xlsm")
    ws = wb["MAPA ALMOX"]
    barcode_data: list[str] = []
    value: str = ""
    match option:
        # Filter by address letter
        case 1:
            letter: str = input(str("Digite a letra do endereco: "))
            for i, cell in enumerate(ws["F"]):
                if i == 0:
                    continue
                value = cell.value
                if value is not None:
                    if value.split("/")[1] == letter:
                        barcode_data.append(str(value))

            return barcode_data
        # Filter by code
        case 2:
            code: str = input(str("Digite o codigo do material: "))
            for cell in ws["F"]:
                value = cell.value
                if value is not None:
                    if value.split("/")[0] == code:
                        barcode_data.append(str(value))
            return barcode_data

        # Get all data without filter
        case 3:
            for i, cell in enumerate(ws["F"]):
                if i == 0:
                    continue
                value = cell.value
                if value is not None:
                    barcode_data.append(str(value))

            return barcode_data

def main():
    """
    Main function to run the program in terminal
    """
    table = Table(title="Opcoes de Impressão")
    table.add_column("[bold underline cyan]Opção", width=5, justify="center")
    table.add_column("[bold underline cyan]Descrição")
    table.add_row("[blue]1", "Por Endereco")
    table.add_row("[blue]2", "Por Codigo do Material")
    table.add_row("[blue]3", "Planilha inteira")
    pt(table)

    option: int = int(input(str("Escolha uma opção: ")))
    barcode_data: list[str] = get_barcode_data(option)
    generate_pdf(barcode_data)

if __name__ == "__main__":
    main()
